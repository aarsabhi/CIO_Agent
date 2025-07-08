"""
File parsing utilities for different document types
"""
import os
import tempfile
from typing import List, Dict, Any
import PyPDF2
import docx
import pandas as pd
from openpyxl import load_workbook

class FileParser:
    """Handle parsing of different file types"""
    
    @staticmethod
    def parse_pdf(file_path: str) -> str:
        """Extract text from PDF file"""
        try:
            with open(file_path, 'rb') as file:
                pdf_reader = PyPDF2.PdfReader(file)
                text = ""
                for page in pdf_reader.pages:
                    text += page.extract_text() + "\n"
                return text.strip()
        except Exception as e:
            raise Exception(f"Error parsing PDF: {str(e)}")
    
    @staticmethod
    def parse_docx(file_path: str) -> str:
        """Extract text from DOCX file"""
        try:
            doc = docx.Document(file_path)
            text = ""
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
            return text.strip()
        except Exception as e:
            raise Exception(f"Error parsing DOCX: {str(e)}")
    
    @staticmethod
    def parse_xlsx(file_path: str) -> Dict[str, Any]:
        """Extract data from XLSX file"""
        try:
            workbook = load_workbook(file_path)
            data = {}
            
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                sheet_data = []
                
                for row in sheet.iter_rows(values_only=True):
                    if any(cell is not None for cell in row):
                        sheet_data.append(row)
                
                data[sheet_name] = sheet_data
            
            return data
        except Exception as e:
            raise Exception(f"Error parsing XLSX: {str(e)}")
    
    @staticmethod
    def parse_csv(file_path: str) -> Dict[str, Any]:
        """Extract data from CSV file"""
        try:
            df = pd.read_csv(file_path)
            return {
                "columns": df.columns.tolist(),
                "data": df.to_dict('records'),
                "shape": df.shape,
                "summary": df.describe().to_dict() if df.select_dtypes(include=['number']).shape[1] > 0 else {}
            }
        except Exception as e:
            raise Exception(f"Error parsing CSV: {str(e)}")
    
    @staticmethod
    def parse_txt(file_path: str) -> str:
        """Extract text from TXT file"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                return file.read().strip()
        except Exception as e:
            raise Exception(f"Error parsing TXT: {str(e)}")
    
    @classmethod
    def parse_file(cls, file_path: str, file_type: str) -> Any:
        """Parse file based on its type"""
        parsers = {
            '.pdf': cls.parse_pdf,
            '.docx': cls.parse_docx,
            '.xlsx': cls.parse_xlsx,
            '.csv': cls.parse_csv,
            '.txt': cls.parse_txt
        }
        
        parser = parsers.get(file_type.lower())
        if not parser:
            raise Exception(f"Unsupported file type: {file_type}")
        
        return parser(file_path)

def chunk_text(text: str, chunk_size: int = 1000, overlap: int = 200) -> List[str]:
    """Split text into overlapping chunks"""
    if len(text) <= chunk_size:
        return [text]
    
    chunks = []
    start = 0
    
    while start < len(text):
        end = start + chunk_size
        chunk = text[start:end]
        
        # Try to break at word boundary
        if end < len(text):
            last_space = chunk.rfind(' ')
            if last_space > chunk_size * 0.8:  # Only if we don't lose too much
                chunk = chunk[:last_space]
                end = start + last_space
        
        chunks.append(chunk.strip())
        start = end - overlap
        
        if start >= len(text):
            break
    
    return chunks